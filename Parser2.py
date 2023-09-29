from docxtpl import DocxTemplate
from openpyxl import load_workbook
doc = DocxTemplate("Карточка ф-9арт-ШАБЛОН.docx")
book = load_workbook(filename="/tmp/Список_военнослужащих.xlsx")
sheet = book['Закрепление']
max_row = sheet.max_row
for i in range(2, max_row +1):
    FIO1 = str(sheet['A' + str(i)].value)
    Npist = sheet['B' + str(i)].value
    context = { 'FIO' : FIO1, 'N_pist' : Npist}
    doc.render(context)
    doc.save('Gotovie_kartochki/' + str(sheet['A' + str(i)].value) + ".docx")

